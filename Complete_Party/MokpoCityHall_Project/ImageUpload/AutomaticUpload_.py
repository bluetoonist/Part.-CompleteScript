# Mokpo BIZ  IMAGE UPLOAD 2019 1/11 완성본
# 2018.11.25~ 2019.1.25 에서 있었던 목포시 일자리 사업
# 업무자동화 Script 증
# 이미지 자동 업로드 Script

import time
from selenium import webdriver
from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.alert import Alert
from selenium.common.exceptions import UnexpectedAlertPresentException

chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
# Your chrome Driver path Setup
chrome_driver = ""
driver = webdriver.Chrome(chrome_driver, options=chrome_options)

import os
url = "http://www.mokpo.go.kr/www/operation_guide/member_login?return_url=/www"

#HtmlPage = WbChronme.page_source
soup = BeautifulSoup(driver.page_source,"html.parser")

RegisterURL = []
GetRegister = []

for rd in soup.find_all("a"):
    link = rd.get("href")
    RegisterURL.append(str(link))

for rd in RegisterURL:
    if "registration" in rd:
        GetRegister.append(rd)

MokpoCityHallURL = "http://www.mokpo.go.kr"

print("IMAGE URL LINK!")
print("이미지 업로드 URL \n",MokpoCityHallURL+GetRegister[0])

NameSearch = driver.page_source
soup = BeautifulSoup(NameSearch,"html.parser")

idx = GetRegister[0].split("=")[2].split("&")[0]
Name_1 = soup.find("div",{"class":"store_search_store"})
Name_2 = Name_1.find("div",{"class":"store_list_wrap"})
Name_3 = Name_2.find("div",{"class":"biz_title_box"})
Name_4 = Name_3.find("strong",{"class":"store_name"})

StoreName =  Name_4.text[2::]

Address = soup.find("div",{"class":"store_depth1"})
Address2 = Address.find("span",{"class":"address store_cont"}).text

Number = soup.find("div",{"class":"store_depth1"})
Number2 = Number.find("span",{"class":"tel store_cont"}).text


import openpyxl
path = 'C:\\Users\\MokpoBIZ\\Desktop\\Working\\Working1\\Excel\\CopyCopy\\copy.xlsx'

wb = openpyxl.load_workbook(path)
ws = wb.active

MaxValue =int()

for r in ws.rows:
    CurrentCounter = r[0].row
    MaxValue = CurrentCounter

print(MaxValue)
SaveValue = MaxValue+1
wb.close()

for r2 in ws.rows:
    if r2[0].row == MaxValue:
        ws.cell(row=SaveValue,column=3).value=StoreName.strip(' ')
        ws.cell(row=SaveValue,column=4).value=Address2.strip(' ')
        ws.cell(row=SaveValue,column=5).value=Number2
    else:
        pass
wb.save(path)
wb.close()


TO_Buiz = "http://www.mokpo.go.kr/business/biz_manager/biz?" \
"idx="+str(idx)+"+search_type=title" \
"&search_word="+StoreName+"" \
"&mode=modify"

ModifyedURL ="http://www.mokpo.go.kr/business/biz_manager/biz?mode=modify&sub_mode=write_01&idx="+idx

count = 1
from selenium.webdriver.common.keys import Keys

driver.get(MokpoCityHallURL + GetRegister[0])
time.sleep(0.5)

from selenium.webdriver.common.alert import Alert
# Image Path setup
BaseDIR = ""
# Image upload Logic
for ld in os.listdir(BaseDIR):
    driver.find_element_by_class_name("add_photo").send_keys(Keys.ENTER)

    try:
        driver.find_element_by_xpath('//*[@id="photo_0"]').send_keys(BaseDIR + ld)
        time.sleep(0.2)
        driver.find_element_by_xpath('//*[@id="upload"]').send_keys('\n')
        
        time.sleep(1)
        Alert(driver).accept()
        
        time.sleep(1.5)
        Alert(driver).accept()

    except UnexpectedAlertPresentException:
        Alert(driver).accept()
    except:
        Alert(driver).accept()

idx_TO = idx+"'"

#대표 이미지 지정으로 넘어감
print("="*50)
driver.execute_script("window.open('http://www.mokpo.go.kr/business/biz_manager/biz?mode=modify&sub_mode=write_01&idx="+idx_TO+");")
#상가 검색을 위한 BIZ 페이지로 넘어감
driver.get('http://www.mokpo.go.kr/biz')
